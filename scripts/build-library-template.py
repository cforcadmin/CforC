# -*- coding: utf-8 -*-
"""
Χτίζει το πρότυπο καταγραφής της Ψηφιακής Βιβλιοθήκης από το taxonomy.json,
που παράγει με τη σειρά του το build-library-template.mjs από το
lib/memberTaxonomy.ts. Καμία λίστα δεν γράφεται με το χέρι.

Τρέξιμο:
  npx tsx scripts/build-library-template.mjs /tmp/taxonomy.json
  python3 scripts/build-library-template.py /tmp/taxonomy.json <φάκελος εξόδου>
"""
import json, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

tx = json.load(open(sys.argv[1], encoding='utf-8'))
out = os.path.join(sys.argv[2], 'Πρότυπο καταγραφής_Digital Library_CforC.xlsx')

wb = Workbook()
CORAL = PatternFill('solid', fgColor='FF8B6A')
GREY = PatternFill('solid', fgColor='EFEFEF')
SAMPLE = PatternFill('solid', fgColor='FFF3EF')
HEAD = Font(bold=True, color='FFFFFF', size=11)
BOLD = Font(bold=True)
ITAL = Font(italic=True, color='8A8A8A')

# ── 1. Καταγραφή ────────────────────────────────────────────────
ws = wb.active
ws.title = 'Καταγραφή'
cols = [('Α/Α', 6), ('Τίτλος', 46), ('Περιγραφή', 70), ('Έτος κυκλοφορίας', 15),
        ('Θεματική', 34), ('Υποθεματική', 34), ('Είδος αρχείου', 44),
        ('Σύνδεσμος πηγής (εκδότη)', 42), ('Σύνδεσμος αρχείου (Drive)', 42), ('Γλώσσα', 12),
        # Συμπληρώνονται αυτόματα όταν η καταχώρηση γίνεται από την ιστοσελίδα.
        # Ο «Κωδικός» εμποδίζει να ξαναγραφτεί η ίδια γραμμή.
        ('Καταχώρηση από', 22), ('Ημερομηνία', 13), ('Κωδικός', 26)]
for i, (name, w) in enumerate(cols, 1):
    c = ws.cell(1, i, name)
    c.fill, c.font = CORAL, HEAD
    c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = 'A2'

# Γραμμή 2: συμπληρωμένο παράδειγμα — το δείγμα της ομάδας, διορθωμένο
sample = [
    1,
    'CREATIVE FLIP: Final Study – Towards more resilient cultural and creative ecosystems',
    'Η μελέτη εξετάζει τους τρόπους αξιολόγησης και ενίσχυσης της ανθεκτικότητας του '
    'πολιτιστικού και δημιουργικού τομέα στην Ευρώπη. Αναλύει τις βασικές δομικές '
    'προκλήσεις του κλάδου, όπως τη χρηματοδότηση, τις συνθήκες εργασίας, τα δικαιώματα '
    'πνευματικής ιδιοκτησίας στην εποχή της Τεχνητής Νοημοσύνης και τις διατομεακές '
    'συνεργασίες. Παράλληλα, παρουσιάζει τα αποτελέσματα του έργου Creative FLIP και '
    'προτείνει συγκεκριμένες πολιτικές κατευθύνσεις για τη μετάβαση από βραχυπρόθεσμα '
    'έργα σε ένα μακροπρόθεσμα βιώσιμο και συνεκτικό οικοσύστημα.',
    2026,
    'Ανθρωπιστικές & Κοινωνικές Επιστήμες',
    'Έρευνα, Πολιτιστική Πολιτική',
    'Report & Analysis (Έκθεση / Τεχνική Αναφορά / Ανάλυση)',
    'https://creativesunite.eu/article/flip-final-study-the-precariat-with-a-portfolio',
    '',
    'Αγγλικά',
    '', '', '',
]
for i, v in enumerate(sample, 1):
    c = ws.cell(2, i, v)
    c.fill = SAMPLE
    c.alignment = Alignment(vertical='top', wrap_text=(i == 3))
ws.row_dimensions[2].height = 92
n = ws.cell(3, 2, '↑ γραμμή‑παράδειγμα: σβήστε την ή γράψτε από κάτω')
n.font = ITAL

for r in range(4, 304):
    ws.cell(r, 1, r - 2)

# ── 2. Λίστες ───────────────────────────────────────────────────
ls = wb.create_sheet('Λίστες')
ls['A1'] = 'ΘΕΜΑΤΙΚΕΣ'
for i, t in enumerate(tx['themes'], 2):
    ls.cell(i, 1, t)
ls['C1'], ls['D1'] = 'ΘΕΜΑΤΙΚΗ', 'ΥΠΟΘΕΜΑΤΙΚΗ'
for i, (th, sub) in enumerate(tx['pairs'], 2):
    ls.cell(i, 3, th)
    ls.cell(i, 4, sub)
ls['F1'] = 'ΕΙΔΗ ΑΡΧΕΙΩΝ'
for i, d in enumerate(tx['docTypes'], 2):
    ls.cell(i, 6, d)
ls['H1'] = 'ΓΛΩΣΣΑ'
for i, g in enumerate(['Ελληνικά', 'Αγγλικά', 'Δίγλωσσο', 'Άλλη'], 2):
    ls.cell(i, 8, g)
for cell in ('A1', 'C1', 'D1', 'F1', 'H1'):
    ls[cell].fill, ls[cell].font = GREY, BOLD
for col, w in (('A', 40), ('C', 34), ('D', 40), ('F', 60), ('H', 14)):
    ls.column_dimensions[col].width = w
ls.auto_filter.ref = f"C1:D{len(tx['pairs']) + 1}"
ls.freeze_panes = 'A2'

# ── αναδυόμενες λίστες ──────────────────────────────────────────
def dv(formula, rng, msg):
    v = DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
    v.error, v.errorTitle = msg, 'Μη έγκυρη τιμή'
    ws.add_data_validation(v)
    v.add(rng)

dv(f"=Λίστες!$A$2:$A${len(tx['themes']) + 1}", 'E2:E303', 'Διάλεξε θεματική από τη λίστα.')
dv(f"=Λίστες!$F$2:$F${len(tx['docTypes']) + 1}", 'G2:G303', 'Διάλεξε είδος αρχείου από τη λίστα.')
dv('=Λίστες!$H$2:$H$5', 'J2:J303', 'Διάλεξε γλώσσα από τη λίστα.')

# ── 3. Οδηγίες ──────────────────────────────────────────────────
od = wb.create_sheet('Οδηγίες')
od.column_dimensions['A'].width = 116
lines = [
    ('ΠΡΟΤΥΠΟ ΚΑΤΑΓΡΑΦΗΣ — CforC Digital Library', True),
    ('', False),
    ('Οι λίστες στο φύλλο «Λίστες» παράγονται αυτόματα από την ταξινομία της ιστοσελίδας — την ίδια', False),
    ('που χρησιμοποιούν τα προφίλ των μελών. Μην τις πληκτρολογείτε ξανά και μην τις αλλάζετε εδώ:', False),
    ('αν χρειάζεται προσθήκη, ζητήστε την και ενημερώνονται μαζί ιστοσελίδα και πρότυπο.', False),
    ('', False),
    ('ΘΕΜΑΤΙΚΗ — μία, από την αναδυόμενη λίστα.', True),
    ('ΥΠΟΘΕΜΑΤΙΚΗ — μία ή περισσότερες, χωρισμένες με κόμμα. Στις στήλες C:D του φύλλου «Λίστες»', True),
    ('   φιλτράρετε με τη θεματική για να δείτε ποιες επιτρέπονται. Στην ιστοσελίδα η υποθεματική', False),
    ('   ακολουθεί αυτόματα τη θεματική — δεν χρειάζεται να λυθεί μέσα στο υπολογιστικό φύλλο.', False),
    ('ΕΤΟΣ — σκέτος αριθμός (π.χ. 2026), όχι κείμενο, όχι δεκαδικά.', True),
    ('', False),
    ('ΣΥΝΔΕΣΜΟΣ ΠΗΓΗΣ — η σελίδα του εκδότη. Είναι ο κύριος σύνδεσμος του αρχείου: σωστή απόδοση,', True),
    ('   χωρίς ζήτημα πνευματικών δικαιωμάτων. Συμπληρώνεται πάντα όταν υπάρχει.', False),
    ('ΣΥΝΔΕΣΜΟΣ ΑΡΧΕΙΟΥ — το αντίγραφο στον φάκελο Assets_documents, ως αρχείο ασφαλείας για όταν', True),
    ('   ο σύνδεσμος του εκδότη πάψει να δουλεύει (συμβαίνει συχνά σε ευρωπαϊκά έργα).', False),
    ('', False),
    ('ΔΙΚΑΙΩΜΑΤΑ ΠΡΟΣΒΑΣΗΣ — ανεβάζετε ΝΕΟ αντίγραφο στον φάκελο· μη μετακινείτε αρχείο που', True),
    ('   υπάρχει ήδη αλλού στο Drive, γιατί κουβαλά μαζί του τον παλιό του διαμοιρασμό.', False),
    ('   Κάθε αρχείο μένει σε «Περιορισμένη πρόσβαση». Αρχείο με «Οποιοσδήποτε έχει τον σύνδεσμο»', False),
    ('   είναι δημόσιο σε όλο το διαδίκτυο, ανεξάρτητα από τα δικαιώματα του φακέλου.', False),
    ('   Την πρόσβαση των μελών την αναλαμβάνει η ιστοσελίδα, που σερβίρει το αρχείο μόνο σε', False),
    ('   συνδεδεμένο μέλος.', False),
    ('', False),
    ('Η γραμμή 2 είναι συμπληρωμένο παράδειγμα. Σβήστε την ή συνεχίστε από κάτω.', True),
    ('', False),
    ('ΟΙ ΤΡΕΙΣ ΤΕΛΕΥΤΑΙΕΣ ΣΤΗΛΕΣ συμπληρώνονται μόνες τους όταν κάποιο μέλος καταχωρεί από την', True),
    ('   ιστοσελίδα. Μην τις πειράζετε — ο «Κωδικός» είναι αυτό που εμποδίζει να γραφτεί η ίδια', False),
    ('   καταχώρηση δύο φορές.', False),
]
for i, (t, b) in enumerate(lines, 1):
    c = od.cell(i, 1, t)
    if b:
        c.font = BOLD

wb.save(out)
print('γράφτηκε:', out)
